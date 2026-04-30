"""Helpers that keep the Streamlit UI thin and the framework reusable."""

from __future__ import annotations

from dataclasses import dataclass, field, replace
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
from pandas import CategoricalDtype

from .config import (
    AdvancedImputationConfig,
    ArtifactConfig,
    CalibrationConfig,
    CleaningConfig,
    ColumnRole,
    ColumnSpec,
    ComparisonConfig,
    CreditRiskDiagnosticConfig,
    CrossValidationConfig,
    DataStructure,
    DiagnosticConfig,
    DocumentationConfig,
    ExecutionConfig,
    ExecutionMode,
    ExplainabilityConfig,
    FeatureDictionaryConfig,
    FeatureDictionaryEntry,
    FeatureEngineeringConfig,
    FeaturePolicyConfig,
    FeatureReviewDecision,
    FeatureReviewDecisionType,
    FeatureSubsetSearchConfig,
    FrameworkConfig,
    ImputationSensitivityConfig,
    ManualReviewConfig,
    MissingValuePolicy,
    ModelConfig,
    ModelType,
    PerformanceConfig,
    PresetName,
    RegulatoryReportConfig,
    ReproducibilityConfig,
    RobustnessConfig,
    ScenarioConfig,
    ScenarioFeatureShock,
    ScenarioShockOperation,
    ScenarioTestConfig,
    SchemaConfig,
    ScorecardBinOverride,
    ScorecardConfig,
    ScorecardWorkbenchConfig,
    SplitConfig,
    SuitabilityCheckConfig,
    TargetConfig,
    TargetMode,
    TransformationConfig,
    TransformationSpec,
    TransformationType,
    VariableSelectionConfig,
    WorkflowGuardrailConfig,
    feature_subset_search_model_types_for_target_mode,
)
from .presets import get_preset_definition, list_preset_definitions

EDITOR_COLUMNS = [
    "enabled",
    "source_name",
    "name",
    "role",
    "dtype",
    "missing_value_policy",
    "missing_value_fill_value",
    "missing_value_group_columns",
    "create_missing_indicator",
    "create_if_missing",
    "default_value",
    "keep_source",
]
FEATURE_DICTIONARY_COLUMNS = [
    "enabled",
    "feature_name",
    "business_name",
    "definition",
    "source_system",
    "unit",
    "allowed_range",
    "missingness_meaning",
    "expected_sign",
    "inclusion_rationale",
    "notes",
]
TRANSFORMATION_EDITOR_COLUMNS = [
    "enabled",
    "transform_type",
    "source_feature",
    "secondary_feature",
    "categorical_value",
    "output_feature",
    "lower_quantile",
    "upper_quantile",
    "parameter_value",
    "window_size",
    "lag_periods",
    "bin_edges",
    "generated_automatically",
    "notes",
]
FEATURE_REVIEW_COLUMNS = [
    "feature_name",
    "decision",
    "rationale",
]
SCORECARD_OVERRIDE_COLUMNS = [
    "feature_name",
    "bin_edges",
    "rationale",
]
SUPPORTED_DTYPES = ["auto", "string", "category", "float", "int", "bool", "datetime"]
SUPPORTED_MISSING_VALUE_POLICIES = [policy.value for policy in MissingValuePolicy]
SUPPORTED_TRANSFORMATION_TYPES = [transform_type.value for transform_type in TransformationType]
SUPPORTED_FEATURE_REVIEW_DECISIONS = [decision.value for decision in FeatureReviewDecisionType]


@dataclass(slots=True)
class GUIBuildInputs:
    """Captures the non-tabular settings collected by the GUI."""

    preset_name: PresetName | None = None
    model: ModelConfig = field(default_factory=ModelConfig)
    cleaning: CleaningConfig = field(default_factory=CleaningConfig)
    feature_engineering: FeatureEngineeringConfig = field(default_factory=FeatureEngineeringConfig)
    comparison: ComparisonConfig = field(default_factory=ComparisonConfig)
    subset_search: FeatureSubsetSearchConfig = field(default_factory=FeatureSubsetSearchConfig)
    feature_policy: FeaturePolicyConfig = field(default_factory=FeaturePolicyConfig)
    feature_dictionary: FeatureDictionaryConfig = field(default_factory=FeatureDictionaryConfig)
    advanced_imputation: AdvancedImputationConfig = field(default_factory=AdvancedImputationConfig)
    transformations: TransformationConfig = field(default_factory=TransformationConfig)
    manual_review: ManualReviewConfig = field(default_factory=ManualReviewConfig)
    suitability_checks: SuitabilityCheckConfig = field(default_factory=SuitabilityCheckConfig)
    workflow_guardrails: WorkflowGuardrailConfig = field(default_factory=WorkflowGuardrailConfig)
    explainability: ExplainabilityConfig = field(default_factory=ExplainabilityConfig)
    calibration: CalibrationConfig = field(default_factory=CalibrationConfig)
    scorecard: ScorecardConfig = field(default_factory=ScorecardConfig)
    scorecard_workbench: ScorecardWorkbenchConfig = field(default_factory=ScorecardWorkbenchConfig)
    imputation_sensitivity: ImputationSensitivityConfig = field(
        default_factory=ImputationSensitivityConfig
    )
    variable_selection: VariableSelectionConfig = field(default_factory=VariableSelectionConfig)
    documentation: DocumentationConfig = field(default_factory=DocumentationConfig)
    regulatory_reporting: RegulatoryReportConfig = field(default_factory=RegulatoryReportConfig)
    scenario_testing: ScenarioTestConfig = field(default_factory=ScenarioTestConfig)
    diagnostics: DiagnosticConfig = field(default_factory=DiagnosticConfig)
    credit_risk: CreditRiskDiagnosticConfig = field(default_factory=CreditRiskDiagnosticConfig)
    robustness: RobustnessConfig = field(default_factory=RobustnessConfig)
    cross_validation: CrossValidationConfig = field(default_factory=CrossValidationConfig)
    reproducibility: ReproducibilityConfig = field(default_factory=ReproducibilityConfig)
    performance: PerformanceConfig = field(default_factory=PerformanceConfig)
    data_structure: DataStructure = DataStructure.CROSS_SECTIONAL
    train_size: float = 0.6
    validation_size: float = 0.2
    test_size: float = 0.2
    random_state: int = 42
    stratify: bool = True
    execution_mode: ExecutionMode = ExecutionMode.FIT_NEW_MODEL
    existing_model_path: Path | None = None
    existing_config_path: Path | None = None
    target_mode: TargetMode = TargetMode.BINARY
    target_output_column: str = "default_flag"
    positive_values_text: str = ""
    drop_target_source_column: bool = False
    pass_through_unconfigured_columns: bool = True
    output_root: Path = Path("artifacts")
    artifacts: ArtifactConfig = field(default_factory=ArtifactConfig)


def build_column_editor_frame(
    dataframe: pd.DataFrame,
    *,
    use_column_name_hints: bool = False,
) -> pd.DataFrame:
    """Creates the editable schema table shown in the GUI."""

    rows: list[dict[str, Any]] = []
    identifier_assigned = False
    for column in dataframe.columns:
        role = ColumnRole.FEATURE
        enabled = True
        if use_column_name_hints:
            role, enabled = infer_column_role_from_name(column)
            if role == ColumnRole.IDENTIFIER:
                if identifier_assigned:
                    role = ColumnRole.IGNORE
                    enabled = False
                else:
                    identifier_assigned = True
        rows.append(
            {
                "enabled": enabled,
                "source_name": column,
                "name": column,
                "role": role.value,
                "dtype": infer_dtype_label(dataframe[column]),
                "missing_value_policy": MissingValuePolicy.INHERIT_DEFAULT.value,
                "missing_value_fill_value": "",
                "missing_value_group_columns": "",
                "create_missing_indicator": False,
                "create_if_missing": False,
                "default_value": "",
                "keep_source": False,
            }
        )

    return pd.DataFrame(rows, columns=EDITOR_COLUMNS)


def infer_column_role_from_name(column_name: str) -> tuple[ColumnRole, bool]:
    """Infers common demo-data column roles from conservative naming hints."""

    normalized = column_name.strip().lower()
    if normalized in {"default_status", "default_flag", "target", "event_flag"}:
        return ColumnRole.TARGET_SOURCE, True
    if normalized in {"as_of_date", "observation_date", "reporting_date", "period_end_date"}:
        return ColumnRole.DATE, True
    if normalized in {"loan_id", "account_id", "facility_id", "entity_id"}:
        return ColumnRole.IDENTIFIER, True
    if normalized in {"legacy_text_field", "drop_me", "placeholder_text"}:
        return ColumnRole.IGNORE, False
    return ColumnRole.FEATURE, True


def build_column_editor_frame_from_schema(schema: SchemaConfig) -> pd.DataFrame:
    """Builds the schema editor from a resolved schema config."""

    rows = [
        {
            "enabled": spec.enabled,
            "source_name": spec.source_name or spec.name,
            "name": spec.name,
            "role": spec.role.value,
            "dtype": spec.dtype or "auto",
            "missing_value_policy": spec.missing_value_policy.value,
            "missing_value_fill_value": (
                "" if spec.missing_value_fill_value is None else spec.missing_value_fill_value
            ),
            "missing_value_group_columns": ", ".join(spec.missing_value_group_columns),
            "create_missing_indicator": spec.create_missing_indicator,
            "create_if_missing": spec.create_if_missing,
            "default_value": "" if spec.default_value is None else spec.default_value,
            "keep_source": spec.keep_source,
        }
        for spec in schema.column_specs
    ]
    return pd.DataFrame(rows, columns=EDITOR_COLUMNS)


def build_feature_dictionary_editor_frame(dataframe: pd.DataFrame) -> pd.DataFrame:
    """Creates a business-metadata table for the current dataframe columns."""

    rows = [
        {
            "enabled": True,
            "feature_name": column_name,
            "business_name": "",
            "definition": "",
            "source_system": "",
            "unit": "",
            "allowed_range": "",
            "missingness_meaning": "",
            "expected_sign": "",
            "inclusion_rationale": "",
            "notes": "",
        }
        for column_name in dataframe.columns
    ]
    return pd.DataFrame(rows, columns=FEATURE_DICTIONARY_COLUMNS)


def build_transformation_editor_frame() -> pd.DataFrame:
    """Creates the governed-transformation editor shown in the GUI."""

    return pd.DataFrame(columns=TRANSFORMATION_EDITOR_COLUMNS)


def build_feature_review_editor_frame() -> pd.DataFrame:
    """Creates the manual feature-review editor shown in the GUI."""

    return pd.DataFrame(columns=FEATURE_REVIEW_COLUMNS)


def build_scorecard_override_editor_frame() -> pd.DataFrame:
    """Creates the scorecard override editor shown in the GUI."""

    return pd.DataFrame(columns=SCORECARD_OVERRIDE_COLUMNS)


def frames_equivalent(left: pd.DataFrame, right: pd.DataFrame) -> bool:
    """Compares editor frames by value while ignoring dtype-only differences."""

    if left.shape != right.shape:
        return False
    if list(left.columns) != list(right.columns):
        return False
    return _frame_records_for_compare(left) == _frame_records_for_compare(right)


def infer_dtype_label(series: pd.Series) -> str:
    """Maps pandas dtypes into the small set of dtypes the framework exposes."""

    if pd.api.types.is_datetime64_any_dtype(series):
        return "datetime"
    if pd.api.types.is_bool_dtype(series):
        return "bool"
    if pd.api.types.is_integer_dtype(series):
        return "int"
    if pd.api.types.is_float_dtype(series):
        return "float"
    if isinstance(series.dtype, CategoricalDtype):
        return "category"
    return "string"


def build_framework_config_from_editor(
    editor_frame: pd.DataFrame,
    inputs: GUIBuildInputs,
    *,
    validate: bool = True,
) -> FrameworkConfig:
    """Converts GUI inputs into the core framework configuration."""

    normalized_editor = normalize_editor_frame(editor_frame)
    column_specs = build_column_specs_from_editor(normalized_editor)
    using_prior_run_config = (
        inputs.execution_mode == ExecutionMode.SCORE_EXISTING_MODEL
        and inputs.existing_config_path is not None
    )
    target_rows = [
        row
        for _, row in normalized_editor.iterrows()
        if str(row["role"]).strip().lower() == ColumnRole.TARGET_SOURCE.value
        and bool(row["enabled"])
    ]
    if not using_prior_run_config and len(target_rows) != 1:
        raise ValueError("Mark exactly one enabled row as the target source in the schema editor.")

    if using_prior_run_config:
        target_source_column = inputs.target_output_column or "target_placeholder"
    else:
        target_row = target_rows[0]
        target_source_column = _clean_text(target_row["name"])
        if not target_source_column:
            raise ValueError("The target-source row must have a valid output column name.")

    date_column = _first_role_column(normalized_editor, ColumnRole.DATE)
    entity_column = _first_role_column(normalized_editor, ColumnRole.IDENTIFIER)

    if (
        not using_prior_run_config
        and inputs.data_structure in {DataStructure.TIME_SERIES, DataStructure.PANEL}
        and not date_column
    ):
        raise ValueError("Time-series and panel runs require one enabled row with role='date'.")
    if (
        not using_prior_run_config
        and inputs.data_structure == DataStructure.PANEL
        and not entity_column
    ):
        raise ValueError("Panel runs require one enabled row with role='identifier'.")

    split_config = SplitConfig(
        data_structure=inputs.data_structure,
        train_size=inputs.train_size,
        validation_size=inputs.validation_size,
        test_size=inputs.test_size,
        random_state=inputs.random_state,
        stratify=inputs.stratify
        if inputs.data_structure == DataStructure.CROSS_SECTIONAL
        else False,
        date_column=date_column,
        entity_column=entity_column,
    )
    split_config.validate()

    config = FrameworkConfig(
        schema=SchemaConfig(
            column_specs=column_specs,
            pass_through_unconfigured_columns=inputs.pass_through_unconfigured_columns,
        ),
        cleaning=inputs.cleaning,
        feature_engineering=inputs.feature_engineering,
        target=TargetConfig(
            source_column=target_source_column,
            mode=inputs.target_mode,
            output_column=inputs.target_output_column,
            positive_values=parse_positive_values(inputs.positive_values_text),
            drop_source_column=inputs.drop_target_source_column,
        ),
        split=split_config,
        preset_name=inputs.preset_name,
        execution=ExecutionConfig(
            mode=inputs.execution_mode,
            existing_model_path=inputs.existing_model_path,
            existing_config_path=inputs.existing_config_path,
        ),
        model=inputs.model,
        comparison=inputs.comparison,
        subset_search=inputs.subset_search,
        feature_policy=inputs.feature_policy,
        feature_dictionary=inputs.feature_dictionary,
        advanced_imputation=inputs.advanced_imputation,
        transformations=inputs.transformations,
        manual_review=inputs.manual_review,
        suitability_checks=inputs.suitability_checks,
        workflow_guardrails=inputs.workflow_guardrails,
        explainability=inputs.explainability,
        calibration=inputs.calibration,
        scorecard=inputs.scorecard,
        scorecard_workbench=inputs.scorecard_workbench,
        imputation_sensitivity=inputs.imputation_sensitivity,
        variable_selection=inputs.variable_selection,
        documentation=inputs.documentation,
        regulatory_reporting=inputs.regulatory_reporting,
        scenario_testing=inputs.scenario_testing,
        diagnostics=inputs.diagnostics,
        credit_risk=inputs.credit_risk,
        robustness=inputs.robustness,
        cross_validation=inputs.cross_validation,
        reproducibility=inputs.reproducibility,
        performance=inputs.performance,
        artifacts=replace(inputs.artifacts, output_root=inputs.output_root),
    )
    if validate:
        config.validate()
    return config


def normalize_editor_frame(editor_frame: pd.DataFrame) -> pd.DataFrame:
    """Normalizes the schema editor into a predictable tabular format."""

    working = editor_frame.copy(deep=True)
    for column in EDITOR_COLUMNS:
        if column not in working.columns:
            working[column] = (
                ""
                if column
                not in {
                    "enabled",
                    "create_if_missing",
                    "keep_source",
                    "create_missing_indicator",
                }
                else False
            )

    working = working.loc[:, EDITOR_COLUMNS].fillna("")
    for boolean_column in [
        "enabled",
        "create_if_missing",
        "keep_source",
        "create_missing_indicator",
    ]:
        working[boolean_column] = working[boolean_column].astype(bool)
    return working


def build_column_specs_from_editor(editor_frame: pd.DataFrame) -> list[ColumnSpec]:
    """Builds framework column specs from the editable schema table."""

    specs: list[ColumnSpec] = []
    seen_names: set[str] = set()

    for _, row in editor_frame.iterrows():
        source_name = _clean_text(row["source_name"])
        output_name = _clean_text(row["name"]) or source_name
        if not output_name:
            continue
        if output_name in seen_names:
            raise ValueError(
                f"Duplicate output column name '{output_name}' found in schema editor."
            )

        seen_names.add(output_name)
        dtype = _clean_text(row["dtype"])
        specs.append(
            ColumnSpec(
                name=output_name,
                source_name=source_name,
                enabled=bool(row["enabled"]),
                dtype=None if dtype in {"", "auto"} else dtype,
                role=ColumnRole(str(row["role"]).strip().lower()),
                missing_value_policy=MissingValuePolicy(
                    str(row["missing_value_policy"]).strip().lower()
                    or MissingValuePolicy.INHERIT_DEFAULT.value
                ),
                missing_value_fill_value=_normalize_default_value(row["missing_value_fill_value"]),
                missing_value_group_columns=_parse_text_list(row["missing_value_group_columns"]),
                create_missing_indicator=bool(row["create_missing_indicator"]),
                create_if_missing=bool(row["create_if_missing"]),
                default_value=_normalize_default_value(row["default_value"]),
                keep_source=bool(row["keep_source"]),
            )
        )

    if not specs:
        raise ValueError("The schema editor does not define any usable columns.")
    return specs


def normalize_feature_dictionary_frame(frame: pd.DataFrame) -> pd.DataFrame:
    """Normalizes the feature-dictionary editor into a predictable shape."""

    working = frame.copy(deep=True)
    for column in FEATURE_DICTIONARY_COLUMNS:
        if column not in working.columns:
            working[column] = False if column == "enabled" else ""
    working = working.loc[:, FEATURE_DICTIONARY_COLUMNS].fillna("")
    working["enabled"] = working["enabled"].astype(bool)
    return working


def normalize_transformation_frame(frame: pd.DataFrame) -> pd.DataFrame:
    """Normalizes the transformation editor into a predictable shape."""

    working = frame.copy(deep=True)
    for column in TRANSFORMATION_EDITOR_COLUMNS:
        if column not in working.columns:
            working[column] = False if column in {"enabled", "generated_automatically"} else ""
    working = working.loc[:, TRANSFORMATION_EDITOR_COLUMNS].fillna("")
    working["enabled"] = working["enabled"].astype(bool)
    working["generated_automatically"] = working["generated_automatically"].astype(bool)
    return working


def build_subset_search_feature_options(
    schema_frame: pd.DataFrame,
    transformation_frame: pd.DataFrame,
) -> list[str]:
    """Returns feature-subset-search options including governed transformation outputs."""

    schema_features = (
        schema_frame.loc[
            schema_frame["enabled"]
            & (
                schema_frame["role"].astype(str).str.strip().str.lower()
                == ColumnRole.FEATURE.value
            ),
            "name",
        ]
        .astype(str)
        .str.strip()
        .replace("", pd.NA)
        .dropna()
        .tolist()
    )
    transformed_features = _resolve_subset_search_transformation_outputs(
        transformation_frame,
    )
    return list(dict.fromkeys([*schema_features, *transformed_features]))


def normalize_feature_review_frame(frame: pd.DataFrame) -> pd.DataFrame:
    """Normalizes the manual feature-review editor into a predictable shape."""

    working = frame.copy(deep=True)
    for column in FEATURE_REVIEW_COLUMNS:
        if column not in working.columns:
            working[column] = ""
    return working.loc[:, FEATURE_REVIEW_COLUMNS].fillna("")


def normalize_scorecard_override_frame(frame: pd.DataFrame) -> pd.DataFrame:
    """Normalizes the scorecard-override editor into a predictable shape."""

    working = frame.copy(deep=True)
    for column in SCORECARD_OVERRIDE_COLUMNS:
        if column not in working.columns:
            working[column] = ""
    return working.loc[:, SCORECARD_OVERRIDE_COLUMNS].fillna("")


def parse_feature_dictionary_frame(frame: pd.DataFrame) -> FeatureDictionaryConfig:
    """Converts the feature-dictionary editor into typed config entries."""

    normalized = normalize_feature_dictionary_frame(frame)
    entries = [
        FeatureDictionaryEntry(
            feature_name=str(row["feature_name"]).strip(),
            business_name=str(row["business_name"]).strip(),
            definition=str(row["definition"]).strip(),
            source_system=str(row["source_system"]).strip(),
            unit=str(row["unit"]).strip(),
            allowed_range=str(row["allowed_range"]).strip(),
            missingness_meaning=str(row["missingness_meaning"]).strip(),
            expected_sign=str(row["expected_sign"]).strip(),
            inclusion_rationale=str(row["inclusion_rationale"]).strip(),
            notes=str(row["notes"]).strip(),
            enabled=bool(row["enabled"]),
        )
        for _, row in normalized.iterrows()
        if str(row["feature_name"]).strip()
    ]
    enabled_entries = [entry for entry in entries if entry.enabled]
    return FeatureDictionaryConfig(
        enabled=bool(enabled_entries),
        entries=entries,
    )


def parse_transformation_frame(frame: pd.DataFrame) -> TransformationConfig:
    """Converts the transformation editor into typed transformation config."""

    normalized = normalize_transformation_frame(frame)
    transformations: list[TransformationSpec] = []
    for _, row in normalized.iterrows():
        if not bool(row["enabled"]):
            continue
        source_feature = str(row["source_feature"]).strip()
        if not source_feature:
            continue
        transformations.append(
            TransformationSpec(
                enabled=True,
                transform_type=TransformationType(str(row["transform_type"]).strip()),
                source_feature=source_feature,
                secondary_feature=_clean_text(row["secondary_feature"]),
                categorical_value=_clean_text(row["categorical_value"]),
                output_feature=_clean_text(row["output_feature"]),
                lower_quantile=_coerce_optional_float(row["lower_quantile"]),
                upper_quantile=_coerce_optional_float(row["upper_quantile"]),
                parameter_value=_coerce_optional_float(row["parameter_value"]),
                window_size=_coerce_optional_int(row["window_size"]),
                lag_periods=_coerce_optional_int(row["lag_periods"]),
                bin_edges=_parse_float_list(row["bin_edges"]),
                generated_automatically=bool(row["generated_automatically"]),
                notes=str(row["notes"]).strip(),
            )
        )
    return TransformationConfig(
        enabled=bool(transformations),
        transformations=transformations,
    )


def parse_manual_review_frames(
    feature_review_frame: pd.DataFrame,
    scorecard_override_frame: pd.DataFrame,
    *,
    reviewer_name: str = "",
    require_review_complete: bool = False,
) -> ManualReviewConfig:
    """Converts review editors into the typed manual-review config."""

    normalized_feature_review = normalize_feature_review_frame(feature_review_frame)
    normalized_scorecard_override = normalize_scorecard_override_frame(scorecard_override_frame)
    feature_decisions = [
        FeatureReviewDecision(
            feature_name=str(row["feature_name"]).strip(),
            decision=FeatureReviewDecisionType(str(row["decision"]).strip()),
            rationale=str(row["rationale"]).strip(),
        )
        for _, row in normalized_feature_review.iterrows()
        if str(row["feature_name"]).strip() and str(row["decision"]).strip()
    ]
    scorecard_bin_overrides = [
        ScorecardBinOverride(
            feature_name=str(row["feature_name"]).strip(),
            bin_edges=_parse_float_list(row["bin_edges"]),
            rationale=str(row["rationale"]).strip(),
        )
        for _, row in normalized_scorecard_override.iterrows()
        if str(row["feature_name"]).strip() and str(row["bin_edges"]).strip()
    ]
    return ManualReviewConfig(
        enabled=bool(feature_decisions or scorecard_bin_overrides),
        reviewer_name=reviewer_name.strip(),
        require_review_complete=require_review_complete,
        feature_decisions=feature_decisions,
        scorecard_bin_overrides=scorecard_bin_overrides,
    )


def parse_positive_values(raw_text: str) -> list[str] | None:
    """Parses comma-separated positive target labels entered by the user."""

    values = [value.strip() for value in raw_text.split(",") if value.strip()]
    return values or None


def _parse_text_list(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    text = str(value).strip()
    if not text:
        return []
    return [item.strip() for item in text.split(",") if item.strip()]


def _first_role_column(editor_frame: pd.DataFrame, role: ColumnRole) -> str | None:
    matching_rows = editor_frame.loc[
        editor_frame["enabled"]
        & (editor_frame["role"].astype(str).str.strip().str.lower() == role.value)
    ]
    if matching_rows.empty:
        return None
    return _clean_text(matching_rows.iloc[0]["name"])


def _clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _normalize_default_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    return value


def _coerce_optional_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        return float(stripped)
    if pd.isna(value):
        return None
    return float(value)


def _coerce_optional_int(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        return int(float(stripped))
    if pd.isna(value):
        return None
    return int(value)


def _frame_records_for_compare(frame: pd.DataFrame) -> list[tuple[Any, ...]]:
    return [
        tuple(_normalize_compare_value(value) for value in row)
        for row in frame.itertuples(index=False, name=None)
    ]


def _normalize_compare_value(value: Any) -> Any:
    if isinstance(value, pd.Timestamp):
        return value.isoformat()
    if hasattr(value, "item") and not isinstance(value, (str, bytes)):
        try:
            value = value.item()
        except Exception:
            pass
    if pd.isna(value):
        return None
    return value


def _parse_float_list(value: Any) -> list[float]:
    if value is None:
        return []
    if isinstance(value, list):
        return [float(item) for item in value]
    if pd.isna(value):
        return []
    text = str(value).strip()
    if not text:
        return []
    return [float(item.strip()) for item in text.split(",") if item.strip()]


def _resolve_subset_search_transformation_outputs(frame: pd.DataFrame) -> list[str]:
    normalized = normalize_transformation_frame(frame)
    output_names: list[str] = []
    for _, row in normalized.iterrows():
        if not bool(row["enabled"]):
            continue
        source_feature = str(row["source_feature"]).strip()
        if not source_feature:
            continue
        transform_name = str(row["transform_type"]).strip()
        if not transform_name:
            continue
        try:
            transform_type = TransformationType(transform_name)
        except ValueError:
            continue
        configured_output = _clean_text(row["output_feature"])
        resolved_output = configured_output or _resolve_transformation_output_name_for_ui(
            transform_type=transform_type,
            source_feature=source_feature,
            secondary_feature=_clean_text(row["secondary_feature"]),
            categorical_value=_clean_text(row["categorical_value"]),
            parameter_value=_coerce_optional_float(row["parameter_value"]),
            window_size=_coerce_optional_int(row["window_size"]),
            lag_periods=_coerce_optional_int(row["lag_periods"]),
        )
        if transform_type == TransformationType.NATURAL_SPLINE:
            spline_df = 4 if _coerce_optional_float(row["parameter_value"]) is None else int(
                _coerce_optional_float(row["parameter_value"])
            )
            output_names.extend(
                f"{resolved_output}_basis_{index + 1}" for index in range(spline_df)
            )
            continue
        if resolved_output:
            output_names.append(resolved_output)
    return output_names


def _resolve_transformation_output_name_for_ui(
    *,
    transform_type: TransformationType,
    source_feature: str,
    secondary_feature: str | None,
    categorical_value: str | None,
    parameter_value: float | None,
    window_size: int | None,
    lag_periods: int | None,
) -> str:
    spec = TransformationSpec(
        transform_type=transform_type,
        source_feature=source_feature,
        secondary_feature=secondary_feature,
        categorical_value=categorical_value,
        parameter_value=parameter_value,
        window_size=window_size,
        lag_periods=lag_periods,
    )
    return TransformationConfig()._resolve_output_name(spec)


def _sanitize_transformation_token(value: str) -> str:
    sanitized = "".join(character if character.isalnum() else "_" for character in value)
    return sanitized.strip("_").lower() or "value"


def build_gui_inputs_from_preset(preset_name: PresetName) -> GUIBuildInputs:
    """Builds a GUI input bundle from a named preset."""

    preset = get_preset_definition(preset_name)
    return GUIBuildInputs(
        preset_name=preset_name,
        model=preset.model,
        feature_engineering=preset.feature_engineering,
        comparison=preset.comparison,
        subset_search=preset.subset_search,
        feature_policy=preset.feature_policy,
        feature_dictionary=preset.feature_dictionary,
        advanced_imputation=preset.advanced_imputation,
        transformations=preset.transformations,
        manual_review=preset.manual_review,
        suitability_checks=preset.suitability_checks,
        workflow_guardrails=preset.workflow_guardrails,
        explainability=preset.explainability,
        calibration=preset.calibration,
        scorecard=preset.scorecard,
        scorecard_workbench=preset.scorecard_workbench,
        imputation_sensitivity=preset.imputation_sensitivity,
        variable_selection=preset.variable_selection,
        documentation=preset.documentation,
        regulatory_reporting=preset.regulatory_reporting,
        scenario_testing=preset.scenario_testing,
        diagnostics=preset.diagnostics,
        credit_risk=preset.credit_risk,
        robustness=preset.robustness,
        cross_validation=preset.cross_validation,
        performance=preset.performance,
        data_structure=preset.data_structure,
        target_mode=preset.target_mode,
        target_output_column=preset.target_output_column,
        positive_values_text=preset.positive_values_text,
    )


def list_gui_preset_options() -> list[tuple[PresetName, str, str]]:
    """Returns preset name, label, and description tuples for the GUI."""

    return [
        (definition.name, definition.label, definition.description)
        for definition in list_preset_definitions()
    ]


def parse_expected_signs(raw_text: str) -> dict[str, str]:
    """Parses feature:direction pairs entered into the GUI."""

    result: dict[str, str] = {}
    for item in raw_text.split(","):
        if ":" not in item:
            continue
        feature_name, direction = item.split(":", 1)
        feature_name = feature_name.strip()
        direction = direction.strip()
        if feature_name and direction:
            result[feature_name] = direction
    return result


def parse_scenario_rows(rows: list[dict[str, Any]]) -> ScenarioTestConfig:
    """Converts a simple editable row payload into typed scenario config objects."""

    grouped_rows: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        scenario_name = str(row.get("scenario_name", "")).strip()
        if not scenario_name:
            continue
        grouped_rows.setdefault(scenario_name, []).append(row)

    scenarios: list[ScenarioConfig] = []
    for scenario_name, scenario_rows in grouped_rows.items():
        shocks: list[ScenarioFeatureShock] = []
        description = ""
        enabled = True
        for row in scenario_rows:
            description = str(row.get("description", description)).strip()
            enabled = bool(row.get("enabled", enabled))
            feature_name = str(row.get("feature_name", "")).strip()
            operation = str(row.get("operation", ScenarioShockOperation.SET.value)).strip()
            value = row.get("value")
            if not feature_name:
                continue
            shocks.append(
                ScenarioFeatureShock(
                    feature_name=feature_name,
                    operation=ScenarioShockOperation(operation),
                    value=value,
                )
            )
        if shocks:
            scenarios.append(
                ScenarioConfig(
                    name=scenario_name,
                    description=description,
                    feature_shocks=shocks,
                    enabled=enabled,
                )
            )

    return ScenarioTestConfig(enabled=bool(scenarios), scenarios=scenarios)


def build_feature_dictionary_frame_from_config(
    feature_dictionary: FeatureDictionaryConfig,
    feature_names: list[str],
) -> pd.DataFrame:
    """Builds an editor-like feature dictionary frame from config entries."""

    entry_map = {entry.feature_name: entry for entry in feature_dictionary.entries}
    rows = []
    for feature_name in list(dict.fromkeys(feature_names)):
        entry = entry_map.get(feature_name)
        rows.append(
            {
                "enabled": True if entry is None else entry.enabled,
                "feature_name": feature_name,
                "business_name": "" if entry is None else entry.business_name,
                "definition": "" if entry is None else entry.definition,
                "source_system": "" if entry is None else entry.source_system,
                "unit": "" if entry is None else entry.unit,
                "allowed_range": "" if entry is None else entry.allowed_range,
                "missingness_meaning": "" if entry is None else entry.missingness_meaning,
                "expected_sign": "" if entry is None else entry.expected_sign,
                "inclusion_rationale": "" if entry is None else entry.inclusion_rationale,
                "notes": "" if entry is None else entry.notes,
            }
        )
    for feature_name, entry in entry_map.items():
        if feature_name in feature_names:
            continue
        rows.append(
            {
                "enabled": entry.enabled,
                "feature_name": feature_name,
                "business_name": entry.business_name,
                "definition": entry.definition,
                "source_system": entry.source_system,
                "unit": entry.unit,
                "allowed_range": entry.allowed_range,
                "missingness_meaning": entry.missingness_meaning,
                "expected_sign": entry.expected_sign,
                "inclusion_rationale": entry.inclusion_rationale,
                "notes": entry.notes,
            }
        )
    return pd.DataFrame(rows, columns=FEATURE_DICTIONARY_COLUMNS)


def build_transformation_frame_from_config(
    transformation_config: TransformationConfig,
) -> pd.DataFrame:
    """Builds an editor-like transformation frame from typed config."""

    rows = [
        {
            "enabled": transformation.enabled,
            "transform_type": transformation.transform_type.value,
            "source_feature": transformation.source_feature,
            "secondary_feature": transformation.secondary_feature or "",
            "categorical_value": transformation.categorical_value or "",
            "output_feature": transformation.output_feature or "",
            "lower_quantile": (
                transformation.lower_quantile if transformation.lower_quantile is not None else ""
            ),
            "upper_quantile": (
                transformation.upper_quantile if transformation.upper_quantile is not None else ""
            ),
            "parameter_value": (
                transformation.parameter_value if transformation.parameter_value is not None else ""
            ),
            "window_size": (
                transformation.window_size if transformation.window_size is not None else ""
            ),
            "lag_periods": (
                transformation.lag_periods if transformation.lag_periods is not None else ""
            ),
            "bin_edges": ", ".join(str(value) for value in transformation.bin_edges),
            "generated_automatically": transformation.generated_automatically,
            "notes": transformation.notes,
        }
        for transformation in transformation_config.transformations
    ]
    return pd.DataFrame(rows, columns=TRANSFORMATION_EDITOR_COLUMNS)


def build_feature_review_frame_from_config(
    manual_review: ManualReviewConfig,
) -> pd.DataFrame:
    """Builds an editor-like feature-review frame from typed config."""

    rows = [
        {
            "feature_name": decision.feature_name,
            "decision": decision.decision.value,
            "rationale": decision.rationale,
        }
        for decision in manual_review.feature_decisions
    ]
    return pd.DataFrame(rows, columns=FEATURE_REVIEW_COLUMNS)


def build_scorecard_override_frame_from_config(
    manual_review: ManualReviewConfig,
) -> pd.DataFrame:
    """Builds an editor-like scorecard override frame from typed config."""

    rows = [
        {
            "feature_name": override.feature_name,
            "bin_edges": ", ".join(str(value) for value in override.bin_edges),
            "rationale": override.rationale,
        }
        for override in manual_review.scorecard_bin_overrides
    ]
    return pd.DataFrame(rows, columns=SCORECARD_OVERRIDE_COLUMNS)


def build_workbook_instructions_frame() -> pd.DataFrame:
    """Returns detailed user guidance for the offline review workbook."""

    rows = [
        {
            "sheet_name": "schema",
            "editable": "yes",
            "purpose": "Define which input columns are used and how they are interpreted.",
            "what_to_change": (
                "Update enabled, output name, role, dtype, missing-value policy, "
                "imputation fields, create-if-missing, default value, and keep-source."
            ),
            "do_not_change": "Do not rename or delete the header row.",
            "upload_notes": (
                "Exactly one enabled row should be target_source unless scoring with a "
                "prior run configuration. Panel and time-series workflows need a date role; "
                "panel workflows also need an identifier role."
            ),
        },
        {
            "sheet_name": "feature_dictionary",
            "editable": "yes",
            "purpose": "Document business meaning, lineage, expected signs, and rationale.",
            "what_to_change": (
                "Complete definitions, source systems, units, allowed ranges, missingness "
                "meaning, expected signs, and inclusion rationale for material features."
            ),
            "do_not_change": "Keep feature_name aligned to the modeled feature or source column.",
            "upload_notes": (
                "Blank fields are allowed, but completed definitions improve the model "
                "documentation pack and validation review."
            ),
        },
        {
            "sheet_name": "transformations",
            "editable": "yes",
            "purpose": "Stage governed feature transformations that are fit on train only.",
            "what_to_change": (
                "Add one row per transformation with transform_type, source_feature, "
                "optional secondary feature, output_feature, and required parameters."
            ),
            "do_not_change": "Do not use transformations only because a metric improved once.",
            "upload_notes": (
                "Transformations are replayed on validation, test, and scored data. "
                "Rows with enabled = false are ignored."
            ),
        },
        {
            "sheet_name": "feature_review",
            "editable": "yes",
            "purpose": "Record manual approve/reject/force decisions for modeled features.",
            "what_to_change": (
                "Add feature_name, decision, and a review rationale when manual review is "
                "enabled or when a feature needs explicit documentation."
            ),
            "do_not_change": "Use only supported decision values from allowed_values.",
            "upload_notes": (
                "Rejected features are removed before training; forced-include and "
                "forced-exclude decisions override normal screening where supported."
            ),
        },
        {
            "sheet_name": "scorecard_overrides",
            "editable": "yes",
            "purpose": "Provide manual bin-edge overrides for scorecard logistic regression.",
            "what_to_change": (
                "Add feature_name, comma-separated internal bin_edges, and rationale."
            ),
            "do_not_change": "Do not include -inf or inf; provide only internal numeric edges.",
            "upload_notes": (
                "Use for policy breakpoints, sparse-bin fixes, or monotonicity review. "
                "Override rows only matter for scorecard runs."
            ),
        },
        {
            "sheet_name": "allowed_values",
            "editable": "reference",
            "purpose": "Lists accepted dropdown values and their meanings.",
            "what_to_change": "Do not edit unless using it as an offline note sheet.",
            "do_not_change": "Workbook upload ignores this sheet.",
            "upload_notes": "Use these values in editable sheets to avoid upload parsing errors.",
        },
        {
            "sheet_name": "transform_catalog",
            "editable": "reference",
            "purpose": "Explains each available transformation and its parameter fields.",
            "what_to_change": "Do not edit unless using it as an offline note sheet.",
            "do_not_change": "Workbook upload ignores this sheet.",
            "upload_notes": (
                "Use this catalog before adding transformation rows so parameters are "
                "defensible and reproducible."
            ),
        },
        {
            "sheet_name": "examples",
            "editable": "reference",
            "purpose": "Shows realistic example rows for common workflow edits.",
            "what_to_change": "Copy patterns into editable sheets as needed.",
            "do_not_change": "Workbook upload ignores this sheet.",
            "upload_notes": "Examples are illustrative and should be adapted to the dataset.",
        },
        {
            "sheet_name": "required_columns",
            "editable": "reference",
            "purpose": "Documents the columns expected by each editable sheet.",
            "what_to_change": "Do not edit unless using it as an offline note sheet.",
            "do_not_change": "Editable sheet headers must remain compatible with this list.",
            "upload_notes": (
                "Missing headers are recreated where possible, but renamed headers lose data."
            ),
        },
    ]
    return pd.DataFrame(rows)


def build_workbook_allowed_values_frame() -> pd.DataFrame:
    """Returns allowed workbook values and concise interpretations."""

    rows: list[dict[str, str]] = []
    rows.extend(
        {
            "field": "schema.role",
            "allowed_value": role.value,
            "meaning": _schema_role_meaning(role),
            "example": "target_source",
        }
        for role in ColumnRole
    )
    rows.extend(
        {
            "field": "schema.dtype",
            "allowed_value": dtype,
            "meaning": _dtype_meaning(dtype),
            "example": "float",
        }
        for dtype in SUPPORTED_DTYPES
    )
    rows.extend(
        {
            "field": "schema.missing_value_policy",
            "allowed_value": policy.value,
            "meaning": _missing_policy_meaning(policy),
            "example": MissingValuePolicy.INHERIT_DEFAULT.value,
        }
        for policy in MissingValuePolicy
    )
    rows.extend(
        {
            "field": "transformations.transform_type",
            "allowed_value": transform_type.value,
            "meaning": _transformation_meaning(transform_type),
            "example": TransformationType.WINSORIZE.value,
        }
        for transform_type in TransformationType
    )
    rows.extend(
        {
            "field": "feature_review.decision",
            "allowed_value": decision.value,
            "meaning": _feature_review_decision_meaning(decision),
            "example": FeatureReviewDecisionType.APPROVE.value,
        }
        for decision in FeatureReviewDecisionType
    )
    rows.extend(
        {
            "field": "boolean fields",
            "allowed_value": value,
            "meaning": "Accepted true/false value for enabled and flag columns.",
            "example": "TRUE",
        }
        for value in ["TRUE", "FALSE", "true", "false", "1", "0"]
    )
    return pd.DataFrame(rows, columns=["field", "allowed_value", "meaning", "example"])


def build_workbook_transform_catalog_frame() -> pd.DataFrame:
    """Returns a user-facing catalog for governed transformation setup."""

    rows = [
        {
            "category": _transformation_category(transform_type),
            "transform_type": transform_type.value,
            "what_it_does": _transformation_meaning(transform_type),
            "when_to_use": _transformation_use_case(transform_type),
            "key_parameters": _transformation_parameter_guidance(transform_type),
            "output_type": _transformation_output_type(transform_type),
        }
        for transform_type in TransformationType
    ]
    return pd.DataFrame(
        rows,
        columns=[
            "category",
            "transform_type",
            "what_it_does",
            "when_to_use",
            "key_parameters",
            "output_type",
        ],
    )


def build_workbook_examples_frame() -> pd.DataFrame:
    """Returns example workbook rows for common offline edits."""

    rows = [
        {
            "sheet_name": "schema",
            "scenario": "Binary target source",
            "example": (
                "enabled=TRUE, source_name=default_status, name=default_status, "
                "role=target_source, dtype=int, missing_value_policy=inherit_default"
            ),
            "notes": "Use exactly one target_source row for normal fit_new_model runs.",
        },
        {
            "sheet_name": "schema",
            "scenario": "Panel date and identifier",
            "example": (
                "as_of_date -> role=date, dtype=datetime; loan_id -> role=identifier, "
                "dtype=string"
            ),
            "notes": "Required for panel splits and time-aware diagnostics.",
        },
        {
            "sheet_name": "schema",
            "scenario": "Constant imputation",
            "example": (
                "missing_value_policy=constant, missing_value_fill_value=0, "
                "create_missing_indicator=TRUE"
            ),
            "notes": "Use when missingness has business meaning and a fixed fill is defensible.",
        },
        {
            "sheet_name": "feature_dictionary",
            "scenario": "Strong feature definition",
            "example": (
                "feature_name=utilization, business_name=Credit Utilization, "
                "definition=Current balance divided by credit limit, "
                "expected_sign=positive, inclusion_rationale=Known PD risk driver"
            ),
            "notes": "Definitions should explain meaning, not just restate the column name.",
        },
        {
            "sheet_name": "transformations",
            "scenario": "Winsorize outliers",
            "example": (
                "enabled=TRUE, transform_type=winsorize, source_feature=income, "
                "output_feature=income_w, lower_quantile=0.01, upper_quantile=0.99"
            ),
            "notes": "Use train-fit quantiles to limit outlier influence.",
        },
        {
            "sheet_name": "transformations",
            "scenario": "Interaction term",
            "example": (
                "enabled=TRUE, transform_type=interaction, source_feature=utilization, "
                "secondary_feature=delinquency_count, output_feature=util_x_dq"
            ),
            "notes": "Only create interactions with a clear business or statistical rationale.",
        },
        {
            "sheet_name": "transformations",
            "scenario": "Quantile bins for scorecard",
            "example": (
                "enabled=TRUE, transform_type=quantile_bins, source_feature=debt_to_income, "
                "output_feature=dti_qbin, parameter_value=5"
            ),
            "notes": "Creates train-fit equal-frequency bins for continuous scorecard inputs.",
        },
        {
            "sheet_name": "transformations",
            "scenario": "Weight of evidence encoding",
            "example": (
                "enabled=TRUE, transform_type=woe_encoding, source_feature=utilization, "
                "output_feature=utilization_woe, parameter_value=5"
            ),
            "notes": "Requires a binary target and is useful for scorecard logistic regression.",
        },
        {
            "sheet_name": "transformations",
            "scenario": "Rare category collapse",
            "example": (
                "enabled=TRUE, transform_type=rare_category_collapse, source_feature=industry, "
                "output_feature=industry_grouped, parameter_value=0.01"
            ),
            "notes": "Groups sparse levels into Other using train-split frequency.",
        },
        {
            "sheet_name": "transformations",
            "scenario": "Date age feature",
            "example": (
                "enabled=TRUE, transform_type=date_age_months, source_feature=origination_date, "
                "secondary_feature=as_of_date, output_feature=loan_age_months"
            ),
            "notes": "Use secondary_feature when age should be calculated row by row.",
        },
        {
            "sheet_name": "transformations",
            "scenario": "Row missingness signal",
            "example": (
                "enabled=TRUE, transform_type=row_missing_share, "
                "source_feature=revenue, ebitda, debt, output_feature=financial_missing_share"
            ),
            "notes": (
                "Comma-separated source_feature values are supported for row missingness only."
            ),
        },
        {
            "sheet_name": "feature_review",
            "scenario": "Reject leakage",
            "example": (
                "feature_name=post_default_fee, decision=reject, "
                "rationale=Observed after default and would leak outcome information"
            ),
            "notes": "Manual review decisions should be specific and audit-readable.",
        },
        {
            "sheet_name": "scorecard_overrides",
            "scenario": "Manual bin edges",
            "example": (
                "feature_name=utilization, bin_edges=0.20, 0.50, 0.80, "
                "rationale=Policy utilization bands reviewed by credit risk"
            ),
            "notes": "Provide internal numeric edges only; omit -inf and inf.",
        },
    ]
    return pd.DataFrame(rows)


def build_workbook_required_columns_frame() -> pd.DataFrame:
    """Returns required columns for each editable workbook sheet."""

    sheet_columns = {
        "schema": EDITOR_COLUMNS,
        "feature_dictionary": FEATURE_DICTIONARY_COLUMNS,
        "transformations": TRANSFORMATION_EDITOR_COLUMNS,
        "feature_review": FEATURE_REVIEW_COLUMNS,
        "scorecard_overrides": SCORECARD_OVERRIDE_COLUMNS,
    }
    rows: list[dict[str, str]] = []
    for sheet_name, columns in sheet_columns.items():
        for column in columns:
            rows.append(
                {
                    "sheet_name": sheet_name,
                    "required_column": column,
                    "can_be_blank": _required_column_can_be_blank(sheet_name, column),
                    "why_it_matters": _required_column_meaning(sheet_name, column),
                }
            )
    return pd.DataFrame(rows)


def build_template_workbook_bytes(
    *,
    schema_frame: pd.DataFrame,
    feature_dictionary_frame: pd.DataFrame,
    transformation_frame: pd.DataFrame,
    feature_review_frame: pd.DataFrame,
    scorecard_override_frame: pd.DataFrame,
) -> bytes:
    """Serializes the editable governance surfaces into an Excel workbook."""

    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        build_workbook_instructions_frame().to_excel(
            writer, sheet_name="instructions", index=False
        )
        build_workbook_allowed_values_frame().to_excel(
            writer, sheet_name="allowed_values", index=False
        )
        build_workbook_transform_catalog_frame().to_excel(
            writer, sheet_name="transform_catalog", index=False
        )
        build_workbook_examples_frame().to_excel(writer, sheet_name="examples", index=False)
        build_workbook_required_columns_frame().to_excel(
            writer, sheet_name="required_columns", index=False
        )
        normalize_editor_frame(schema_frame).to_excel(writer, sheet_name="schema", index=False)
        normalize_feature_dictionary_frame(feature_dictionary_frame).to_excel(
            writer, sheet_name="feature_dictionary", index=False
        )
        normalize_transformation_frame(transformation_frame).to_excel(
            writer, sheet_name="transformations", index=False
        )
        normalize_feature_review_frame(feature_review_frame).to_excel(
            writer, sheet_name="feature_review", index=False
        )
        normalize_scorecard_override_frame(scorecard_override_frame).to_excel(
            writer, sheet_name="scorecard_overrides", index=False
        )
        _format_template_workbook(writer.book)

    return buffer.getvalue()


def load_template_workbook(source: str | Path | BytesIO | Any) -> dict[str, pd.DataFrame]:
    """Loads an exported workbook template back into editable dataframes."""

    workbook = pd.read_excel(source, sheet_name=None)
    return {
        "schema": normalize_editor_frame(workbook.get("schema", pd.DataFrame())),
        "feature_dictionary": normalize_feature_dictionary_frame(
            workbook.get("feature_dictionary", pd.DataFrame())
        ),
        "transformations": normalize_transformation_frame(
            workbook.get("transformations", pd.DataFrame())
        ),
        "feature_review": normalize_feature_review_frame(
            workbook.get("feature_review", pd.DataFrame())
        ),
        "scorecard_overrides": normalize_scorecard_override_frame(
            workbook.get("scorecard_overrides", pd.DataFrame())
        ),
    }


def _schema_role_meaning(role: ColumnRole) -> str:
    meanings = {
        ColumnRole.FEATURE: "Predictor candidate available to the model.",
        ColumnRole.TARGET_SOURCE: "Raw outcome column used to build the modeled target.",
        ColumnRole.DATE: "Date/time field used for splits and time-aware diagnostics.",
        ColumnRole.IDENTIFIER: "Record/entity ID used for panel structure and traceability.",
        ColumnRole.IGNORE: "Column excluded from modeling and downstream feature treatment.",
    }
    return meanings[role]


def _dtype_meaning(dtype: str) -> str:
    meanings = {
        "auto": "Infer the dtype from the source dataframe.",
        "string": "Treat values as text.",
        "category": "Treat values as categorical levels.",
        "float": "Treat values as decimal numeric values.",
        "int": "Treat values as integer numeric values.",
        "bool": "Treat values as true/false flags.",
        "datetime": "Parse values as dates or timestamps.",
    }
    return meanings.get(dtype, "Supported Quant Studio dtype.")


def _missing_policy_meaning(policy: MissingValuePolicy) -> str:
    meanings = {
        MissingValuePolicy.INHERIT_DEFAULT: (
            "Use the framework default: numeric median and categorical mode."
        ),
        MissingValuePolicy.NONE: "Do not impute this column.",
        MissingValuePolicy.MEAN: "Fill numeric missing values with the train-split mean.",
        MissingValuePolicy.MEDIAN: "Fill numeric missing values with the train-split median.",
        MissingValuePolicy.MODE: "Fill missing values with the train-split most frequent value.",
        MissingValuePolicy.CONSTANT: "Fill missing values with missing_value_fill_value.",
        MissingValuePolicy.FORWARD_FILL: "Carry prior value forward in time-aware data.",
        MissingValuePolicy.BACKWARD_FILL: "Carry next value backward in time-aware data.",
        MissingValuePolicy.KNN: (
            "Use KNN model-based imputation when advanced imputation is enabled."
        ),
        MissingValuePolicy.ITERATIVE: (
            "Use iterative model-based imputation when advanced imputation is enabled."
        ),
    }
    return meanings[policy]


def _transformation_meaning(transform_type: TransformationType) -> str:
    meanings = {
        TransformationType.WINSORIZE: "Clip numeric tails using train-fit quantiles.",
        TransformationType.LOG1P: "Apply log(1 + x) for skewed numeric values above -1.",
        TransformationType.BOX_COX: "Apply a train-fit Box-Cox power transform.",
        TransformationType.NATURAL_SPLINE: "Create natural cubic spline basis features.",
        TransformationType.YEO_JOHNSON: (
            "Apply a power transform that supports non-positive values."
        ),
        TransformationType.CAPPED_ZSCORE: "Standardize then cap extreme z-scores.",
        TransformationType.PIECEWISE_LINEAR: "Create a hinge term above a configured cut point.",
        TransformationType.STANDARD_SCALE: "Subtract the train mean and divide by train std.",
        TransformationType.ROBUST_SCALE: "Subtract the train median and divide by train IQR.",
        TransformationType.MIN_MAX_SCALE: "Scale values to the train minimum/maximum range.",
        TransformationType.PERCENTILE_RANK: "Map values to their train percentile position.",
        TransformationType.NORMAL_SCORE: "Convert percentile rank to a normal-score value.",
        TransformationType.SIGNED_LOG1P: "Apply sign(x) * log(1 + abs(x)).",
        TransformationType.SQRT: "Apply square root to non-negative values.",
        TransformationType.RECIPROCAL: "Create 1 / x with safe zero handling.",
        TransformationType.SQUARE: "Square the numeric feature.",
        TransformationType.POWER: "Raise the numeric feature to parameter_value.",
        TransformationType.ABSOLUTE_VALUE: "Use the absolute value of the numeric feature.",
        TransformationType.CENTER_MEAN: "Subtract the train-split mean.",
        TransformationType.CENTER_MEDIAN: "Subtract the train-split median.",
        TransformationType.RATIO: "Create source_feature divided by secondary_feature.",
        TransformationType.SAFE_RATIO: "Create a ratio with explicit near-zero protection.",
        TransformationType.MARGIN_RATIO: "Create (source - secondary) / abs(source).",
        TransformationType.DEBT_SERVICE_RATIO: (
            "Create source divided by secondary for coverage ratios."
        ),
        TransformationType.ADD: "Add source_feature and secondary_feature.",
        TransformationType.SUBTRACT: "Subtract secondary_feature from source_feature.",
        TransformationType.PRODUCT: "Multiply source_feature by secondary_feature.",
        TransformationType.INTERACTION: "Create source_feature multiplied by secondary_feature.",
        TransformationType.LAG: "Create a lagged version of a time-aware feature.",
        TransformationType.DIFFERENCE: "Create current minus lagged value.",
        TransformationType.EWMA: "Create an exponentially weighted moving average.",
        TransformationType.ROLLING_MEAN: "Create a rolling mean over the configured window.",
        TransformationType.ROLLING_MEDIAN: "Create a rolling median over the configured window.",
        TransformationType.ROLLING_MIN: "Create a rolling minimum over the configured window.",
        TransformationType.ROLLING_MAX: "Create a rolling maximum over the configured window.",
        TransformationType.ROLLING_STD: "Create a rolling standard deviation.",
        TransformationType.ROLLING_SUM: "Create a rolling sum over prior observations.",
        TransformationType.ROLLING_RANGE: "Create rolling max minus rolling min.",
        TransformationType.ROLLING_CV: "Create rolling standard deviation divided by rolling mean.",
        TransformationType.ROLLING_SLOPE: "Estimate a rolling linear trend slope.",
        TransformationType.EXPANDING_MEAN: "Create a prior-observation expanding mean.",
        TransformationType.CUMULATIVE_SUM: "Create a prior-observation cumulative sum.",
        TransformationType.CUMULATIVE_COUNT: "Count prior observations in time order.",
        TransformationType.MONTHS_SINCE_EVENT: (
            "Estimate elapsed months since the source was positive."
        ),
        TransformationType.CHANGE_FROM_BASELINE: "Subtract the first observed value in time order.",
        TransformationType.PCT_CHANGE: "Create percent change over the lag period.",
        TransformationType.MANUAL_BINS: "Create ordered categorical bins from numeric edges.",
        TransformationType.QUANTILE_BINS: "Create train-fit equal-frequency numeric bins.",
        TransformationType.EQUAL_WIDTH_BINS: "Create train-fit equal-width numeric bins.",
        TransformationType.MONOTONIC_BINS: (
            "Create bins whose train target rates are monotonic when possible."
        ),
        TransformationType.WOE_ENCODING: "Encode bins/categories using weight of evidence.",
        TransformationType.BAD_RATE_ENCODING: "Encode bins/categories using train target rate.",
        TransformationType.RARE_CATEGORY_COLLAPSE: "Group sparse category levels into Other.",
        TransformationType.FREQUENCY_ENCODING: "Encode categories by train frequency share.",
        TransformationType.ORDINAL_ENCODING: "Map categories to deterministic integer codes.",
        TransformationType.TARGET_ENCODING: "Encode categories by smoothed train target mean.",
        TransformationType.DATE_YEAR: "Extract calendar year from a date.",
        TransformationType.DATE_MONTH: "Extract calendar month from a date.",
        TransformationType.DATE_QUARTER: "Extract calendar quarter from a date.",
        TransformationType.DATE_MONTH_END_FLAG: "Flag whether a date falls on month end.",
        TransformationType.DATE_FISCAL_QUARTER: (
            "Extract fiscal quarter using parameter_value start month."
        ),
        TransformationType.DATE_AGE_DAYS: "Calculate day age from source date to reference date.",
        TransformationType.DATE_AGE_MONTHS: (
            "Calculate month age from source date to reference date."
        ),
        TransformationType.ROW_MISSING_COUNT: "Count missing values across listed source fields.",
        TransformationType.ROW_MISSING_SHARE: (
            "Calculate missing share across listed source fields."
        ),
        TransformationType.ANY_MISSING_FLAG: "Flag whether any listed source field is missing.",
    }
    return meanings.get(transform_type, "Supported governed transformation.")


def _transformation_category(transform_type: TransformationType) -> str:
    categories = {
        TransformationType.WINSORIZE: "Numeric shape",
        TransformationType.LOG1P: "Numeric shape",
        TransformationType.BOX_COX: "Numeric shape",
        TransformationType.NATURAL_SPLINE: "Numeric shape",
        TransformationType.YEO_JOHNSON: "Numeric shape",
        TransformationType.CAPPED_ZSCORE: "Numeric shape",
        TransformationType.PIECEWISE_LINEAR: "Numeric shape",
        TransformationType.STANDARD_SCALE: "Scaling and rank",
        TransformationType.ROBUST_SCALE: "Scaling and rank",
        TransformationType.MIN_MAX_SCALE: "Scaling and rank",
        TransformationType.PERCENTILE_RANK: "Scaling and rank",
        TransformationType.NORMAL_SCORE: "Scaling and rank",
        TransformationType.SIGNED_LOG1P: "Numeric shape",
        TransformationType.SQRT: "Numeric shape",
        TransformationType.RECIPROCAL: "Numeric shape",
        TransformationType.SQUARE: "Numeric shape",
        TransformationType.POWER: "Numeric shape",
        TransformationType.ABSOLUTE_VALUE: "Numeric shape",
        TransformationType.CENTER_MEAN: "Scaling and rank",
        TransformationType.CENTER_MEDIAN: "Scaling and rank",
        TransformationType.RATIO: "Combinations",
        TransformationType.SAFE_RATIO: "Combinations",
        TransformationType.MARGIN_RATIO: "Combinations",
        TransformationType.DEBT_SERVICE_RATIO: "Combinations",
        TransformationType.ADD: "Combinations",
        TransformationType.SUBTRACT: "Combinations",
        TransformationType.PRODUCT: "Combinations",
        TransformationType.INTERACTION: "Combinations",
        TransformationType.LAG: "Time-aware",
        TransformationType.DIFFERENCE: "Time-aware",
        TransformationType.EWMA: "Time-aware",
        TransformationType.ROLLING_MEAN: "Time-aware",
        TransformationType.ROLLING_MEDIAN: "Time-aware",
        TransformationType.ROLLING_MIN: "Time-aware",
        TransformationType.ROLLING_MAX: "Time-aware",
        TransformationType.ROLLING_STD: "Time-aware",
        TransformationType.ROLLING_SUM: "Time-aware",
        TransformationType.ROLLING_RANGE: "Time-aware",
        TransformationType.ROLLING_CV: "Time-aware",
        TransformationType.ROLLING_SLOPE: "Time-aware",
        TransformationType.EXPANDING_MEAN: "Time-aware",
        TransformationType.CUMULATIVE_SUM: "Time-aware",
        TransformationType.CUMULATIVE_COUNT: "Time-aware",
        TransformationType.MONTHS_SINCE_EVENT: "Time-aware",
        TransformationType.CHANGE_FROM_BASELINE: "Time-aware",
        TransformationType.PCT_CHANGE: "Time-aware",
        TransformationType.MANUAL_BINS: "Binning and encoding",
        TransformationType.QUANTILE_BINS: "Binning and encoding",
        TransformationType.EQUAL_WIDTH_BINS: "Binning and encoding",
        TransformationType.MONOTONIC_BINS: "Binning and encoding",
        TransformationType.WOE_ENCODING: "Binning and encoding",
        TransformationType.BAD_RATE_ENCODING: "Binning and encoding",
        TransformationType.RARE_CATEGORY_COLLAPSE: "Binning and encoding",
        TransformationType.FREQUENCY_ENCODING: "Binning and encoding",
        TransformationType.ORDINAL_ENCODING: "Binning and encoding",
        TransformationType.TARGET_ENCODING: "Binning and encoding",
        TransformationType.DATE_YEAR: "Date features",
        TransformationType.DATE_MONTH: "Date features",
        TransformationType.DATE_QUARTER: "Date features",
        TransformationType.DATE_MONTH_END_FLAG: "Date features",
        TransformationType.DATE_FISCAL_QUARTER: "Date features",
        TransformationType.DATE_AGE_DAYS: "Date features",
        TransformationType.DATE_AGE_MONTHS: "Date features",
        TransformationType.ROW_MISSING_COUNT: "Missingness",
        TransformationType.ROW_MISSING_SHARE: "Missingness",
        TransformationType.ANY_MISSING_FLAG: "Missingness",
    }
    return categories.get(transform_type, "Other")


def _transformation_use_case(transform_type: TransformationType) -> str:
    category = _transformation_category(transform_type)
    if category == "Numeric shape":
        return "Use when scale, skew, tails, or nonlinear marginal effects need treatment."
    if category == "Scaling and rank":
        return "Use for models sensitive to feature magnitude or heavy-tailed ranks."
    if category == "Combinations":
        return "Use when a business ratio, spread, coverage metric, or interaction is expected."
    if category == "Time-aware":
        return "Use for panel or time-series data where prior behavior has predictive value."
    if category == "Binning and encoding":
        return "Use for scorecards, categorical signal, sparse levels, or nonlinear buckets."
    if category == "Date features":
        return "Use to convert governed dates into numeric model inputs."
    if category == "Missingness":
        return "Use when missingness patterns contain business or operational signal."
    return "Use when the transformation has a documented modeling rationale."


def _transformation_parameter_guidance(transform_type: TransformationType) -> str:
    if transform_type == TransformationType.WINSORIZE:
        return "lower_quantile and upper_quantile; defaults 0.01 and 0.99."
    if transform_type == TransformationType.NATURAL_SPLINE:
        return "parameter_value controls degrees of freedom; minimum 3, default 4."
    if transform_type == TransformationType.CAPPED_ZSCORE:
        return "parameter_value controls z cap; default 3."
    if transform_type in {TransformationType.PIECEWISE_LINEAR, TransformationType.POWER}:
        return "parameter_value is required."
    if transform_type in {
        TransformationType.QUANTILE_BINS,
        TransformationType.EQUAL_WIDTH_BINS,
        TransformationType.MONOTONIC_BINS,
        TransformationType.WOE_ENCODING,
        TransformationType.BAD_RATE_ENCODING,
    }:
        return "parameter_value controls bin count; bin_edges can override for WOE/bad-rate."
    if transform_type == TransformationType.MANUAL_BINS:
        return "bin_edges is required; provide internal numeric edges only."
    if transform_type == TransformationType.RARE_CATEGORY_COLLAPSE:
        return "parameter_value is minimum train share; default 0.01."
    if transform_type == TransformationType.TARGET_ENCODING:
        return "parameter_value is smoothing strength; default 20."
    if transform_type in {
        TransformationType.RATIO,
        TransformationType.SAFE_RATIO,
        TransformationType.MARGIN_RATIO,
        TransformationType.DEBT_SERVICE_RATIO,
        TransformationType.ADD,
        TransformationType.SUBTRACT,
        TransformationType.PRODUCT,
        TransformationType.INTERACTION,
    }:
        return "secondary_feature is required; categorical_value is optional for interactions."
    if transform_type in {
        TransformationType.LAG,
        TransformationType.DIFFERENCE,
        TransformationType.PCT_CHANGE,
    }:
        return "lag_periods controls the lookback; default 1."
    if "rolling" in transform_type.value or transform_type == TransformationType.EWMA:
        return "window_size controls the lookback window; default 3."
    if transform_type == TransformationType.DATE_FISCAL_QUARTER:
        return "parameter_value is fiscal year start month; default 1."
    if transform_type in {TransformationType.DATE_AGE_DAYS, TransformationType.DATE_AGE_MONTHS}:
        return "secondary_feature can hold row-level reference date; otherwise split date is used."
    if transform_type in {
        TransformationType.ROW_MISSING_COUNT,
        TransformationType.ROW_MISSING_SHARE,
        TransformationType.ANY_MISSING_FLAG,
    }:
        return "source_feature accepts comma-separated column names."
    return "No additional parameter required beyond source_feature and output_feature."


def _transformation_output_type(transform_type: TransformationType) -> str:
    if transform_type in {
        TransformationType.MANUAL_BINS,
        TransformationType.QUANTILE_BINS,
        TransformationType.EQUAL_WIDTH_BINS,
        TransformationType.MONOTONIC_BINS,
        TransformationType.RARE_CATEGORY_COLLAPSE,
    }:
        return "categorical"
    if transform_type == TransformationType.NATURAL_SPLINE:
        return "multiple numeric columns"
    return "numeric"


def _feature_review_decision_meaning(decision: FeatureReviewDecisionType) -> str:
    meanings = {
        FeatureReviewDecisionType.APPROVE: "Explicitly approve a feature for consideration.",
        FeatureReviewDecisionType.REJECT: "Reject a feature before model training.",
        FeatureReviewDecisionType.FORCE_INCLUDE: (
            "Force a feature to survive selection where allowed."
        ),
        FeatureReviewDecisionType.FORCE_EXCLUDE: (
            "Force a feature out even if screening ranks it well."
        ),
    }
    return meanings[decision]


def _required_column_can_be_blank(sheet_name: str, column: str) -> str:
    required = {
        "schema": {"enabled", "source_name", "name", "role"},
        "feature_dictionary": {"feature_name"},
        "transformations": {"enabled", "transform_type", "source_feature", "output_feature"},
        "feature_review": {"feature_name", "decision"},
        "scorecard_overrides": {"feature_name", "bin_edges"},
    }
    return "no" if column in required.get(sheet_name, set()) else "yes"


def _required_column_meaning(sheet_name: str, column: str) -> str:
    common = {
        "enabled": "Controls whether the row is active.",
        "source_name": "Input column name from the source dataset.",
        "name": "Output column name used by the framework.",
        "role": "How the column is treated by the pipeline.",
        "dtype": "How the column should be parsed.",
        "feature_name": "Feature name the row documents or controls.",
        "rationale": "Audit-readable explanation for the decision.",
        "notes": "Optional reviewer notes.",
    }
    sheet_specific = {
        ("schema", "missing_value_policy"): "Train-fit imputation policy for this column.",
        ("schema", "missing_value_fill_value"): "Constant fill value when policy is constant.",
        ("schema", "missing_value_group_columns"): (
            "Comma-separated segment columns for grouped imputation."
        ),
        ("schema", "create_missing_indicator"): "Creates a model feature flag for missingness.",
        ("schema", "create_if_missing"): "Creates a synthetic column when the source is absent.",
        ("schema", "default_value"): "Default value for create-if-missing columns.",
        ("schema", "keep_source"): "Keeps source column when output name differs.",
        ("feature_dictionary", "business_name"): "Human-readable feature name.",
        ("feature_dictionary", "definition"): "Plain-English definition of the feature.",
        ("feature_dictionary", "source_system"): "System or process that produced the feature.",
        ("feature_dictionary", "unit"): "Measurement unit or scale.",
        ("feature_dictionary", "allowed_range"): "Expected valid range or categories.",
        ("feature_dictionary", "missingness_meaning"): "Business meaning of missing values.",
        ("feature_dictionary", "expected_sign"): "Expected relationship to the target.",
        ("feature_dictionary", "inclusion_rationale"): "Why the feature belongs in development.",
        ("transformations", "transform_type"): "Governed transformation operation.",
        ("transformations", "source_feature"): (
            "Primary input feature, or comma-separated feature list for row missingness."
        ),
        ("transformations", "secondary_feature"): (
            "Second input for arithmetic/interaction transforms or date-age reference date."
        ),
        ("transformations", "categorical_value"): (
            "Category value for category-specific transforms."
        ),
        ("transformations", "output_feature"): "Generated feature name.",
        ("transformations", "lower_quantile"): "Lower quantile for winsorization.",
        ("transformations", "upper_quantile"): "Upper quantile for winsorization.",
        ("transformations", "parameter_value"): (
            "Numeric cut point, cap, bin count, smoothing, or transform parameter."
        ),
        ("transformations", "window_size"): "Rolling or EWMA window.",
        ("transformations", "lag_periods"): "Lag length for time-aware transforms.",
        ("transformations", "bin_edges"): "Comma-separated internal bin edges.",
        ("transformations", "generated_automatically"): "Flags system-suggested rows.",
        ("feature_review", "decision"): "Manual review decision.",
        ("scorecard_overrides", "bin_edges"): "Comma-separated scorecard internal bin edges.",
    }
    return sheet_specific.get((sheet_name, column), common.get(column, "Workbook field."))


def _format_template_workbook(workbook: Any) -> None:
    """Applies formatting, comments, and dropdown validation to the workbook."""

    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    header_fill = PatternFill("solid", fgColor="EAF2FF")
    reference_fill = PatternFill("solid", fgColor="F6F8FB")
    header_font = Font(bold=True, color="10213F")
    guidance_font = Font(color="52657F")

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        worksheet.freeze_panes = "A2"
        worksheet.sheet_view.showGridLines = False
        if worksheet.max_row and worksheet.max_column:
            worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.fill = reference_fill if sheet_name in _REFERENCE_SHEETS else header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if sheet_name in _REFERENCE_SHEETS:
                    cell.font = guidance_font
        for column_index in range(1, worksheet.max_column + 1):
            letter = get_column_letter(column_index)
            values = [
                str(worksheet.cell(row=row_index, column=column_index).value or "")
                for row_index in range(1, min(worksheet.max_row, 80) + 1)
            ]
            worksheet.column_dimensions[letter].width = min(max(map(len, values)) + 2, 58)

    _apply_header_comments(workbook, Comment)
    _apply_workbook_validations(workbook, DataValidation)
    _apply_tab_colors(workbook)


_REFERENCE_SHEETS = {
    "instructions",
    "allowed_values",
    "transform_catalog",
    "examples",
    "required_columns",
}


_HEADER_COMMENTS: dict[str, dict[str, str]] = {
    "schema": {
        "enabled": "TRUE keeps this row active; FALSE ignores it.",
        "source_name": "Must match the input data column unless create_if_missing is TRUE.",
        "name": "Framework output name. Keep stable for model reproducibility.",
        "role": "Use feature, target_source, date, identifier, or ignore.",
        "dtype": "Use auto unless the source needs explicit parsing.",
        "missing_value_policy": "Policy is fit on train and replayed downstream.",
        "missing_value_fill_value": "Only used when missing_value_policy is constant.",
        "missing_value_group_columns": "Optional comma-separated columns for grouped imputation.",
        "create_missing_indicator": "TRUE creates a separate missingness flag feature.",
        "create_if_missing": "TRUE creates this column if the input data does not contain it.",
        "default_value": "Value used for create_if_missing rows.",
        "keep_source": "TRUE keeps the raw source when name differs from source_name.",
    },
    "feature_dictionary": {
        "enabled": "TRUE includes this row in documentation outputs.",
        "feature_name": "Feature or source column being documented.",
        "business_name": "Plain-English name used in documentation.",
        "definition": "Describe how the feature is calculated and what it means.",
        "source_system": "System, feed, or process that produced the feature.",
        "unit": "Currency, percentage, count, date, flag, bucket, etc.",
        "allowed_range": "Expected valid values or range.",
        "missingness_meaning": "Explain whether missing values have business meaning.",
        "expected_sign": (
            "Expected direction, such as positive, negative, increasing, or decreasing."
        ),
        "inclusion_rationale": "Why this feature should be considered for development.",
        "notes": "Additional reviewer or business notes.",
    },
    "transformations": {
        "enabled": "TRUE applies the transformation; FALSE keeps it as a draft.",
        "transform_type": "Choose a supported transform from transform_catalog.",
        "source_feature": (
            "Primary input feature. For row missingness transforms, comma-separated "
            "features are allowed."
        ),
        "secondary_feature": (
            "Required for ratio, arithmetic, and interaction transforms; optional "
            "reference date for date-age transforms."
        ),
        "categorical_value": "Optional category value for category-specific interactions.",
        "output_feature": "Generated output feature name.",
        "lower_quantile": "Winsorization lower quantile, e.g. 0.01.",
        "upper_quantile": "Winsorization upper quantile, e.g. 0.99.",
        "parameter_value": (
            "Cut point, cap, spline df, bin count, smoothing, fiscal start month, "
            "or other numeric parameter."
        ),
        "window_size": "Window length for rolling or EWMA transforms.",
        "lag_periods": "Lag length for lag, difference, and pct_change.",
        "bin_edges": "Internal numeric bin edges such as 0.2, 0.5, 0.8.",
        "generated_automatically": "TRUE means the row was suggested by the system.",
        "notes": "Governance rationale for the transformation.",
    },
    "feature_review": {
        "feature_name": "Feature controlled by the manual decision.",
        "decision": "Use approve, reject, force_include, or force_exclude.",
        "rationale": "Explain the review decision in audit-readable language.",
    },
    "scorecard_overrides": {
        "feature_name": "Numeric feature receiving manual scorecard bins.",
        "bin_edges": "Comma-separated internal bin edges. Do not include -inf or inf.",
        "rationale": "Explain policy, monotonicity, sparse-bin, or business breakpoint logic.",
    },
}


def _apply_header_comments(workbook: Any, comment_class: Any) -> None:
    for sheet_name, comments in _HEADER_COMMENTS.items():
        if sheet_name not in workbook.sheetnames:
            continue
        worksheet = workbook[sheet_name]
        headers = {
            str(cell.value): column_index
            for column_index, cell in enumerate(worksheet[1], start=1)
            if cell.value is not None
        }
        for column_name, comment_text in comments.items():
            column_index = headers.get(column_name)
            if column_index is None:
                continue
            worksheet.cell(row=1, column=column_index).comment = comment_class(
                comment_text,
                "Quant Studio",
            )


def _apply_workbook_validations(workbook: Any, validation_class: Any) -> None:
    _add_dropdown_validation(
        workbook,
        validation_class,
        sheet_name="schema",
        column_name="role",
        values=[role.value for role in ColumnRole],
    )
    _add_dropdown_validation(
        workbook,
        validation_class,
        sheet_name="schema",
        column_name="dtype",
        values=SUPPORTED_DTYPES,
    )
    _add_dropdown_validation(
        workbook,
        validation_class,
        sheet_name="schema",
        column_name="missing_value_policy",
        values=SUPPORTED_MISSING_VALUE_POLICIES,
    )
    _add_dropdown_validation(
        workbook,
        validation_class,
        sheet_name="transformations",
        column_name="transform_type",
        values=SUPPORTED_TRANSFORMATION_TYPES,
    )
    _add_dropdown_validation(
        workbook,
        validation_class,
        sheet_name="feature_review",
        column_name="decision",
        values=SUPPORTED_FEATURE_REVIEW_DECISIONS,
    )
    for sheet_name, column_name in [
        ("schema", "enabled"),
        ("schema", "create_missing_indicator"),
        ("schema", "create_if_missing"),
        ("schema", "keep_source"),
        ("feature_dictionary", "enabled"),
        ("transformations", "enabled"),
        ("transformations", "generated_automatically"),
    ]:
        _add_dropdown_validation(
            workbook,
            validation_class,
            sheet_name=sheet_name,
            column_name=column_name,
            values=["TRUE", "FALSE"],
        )


def _add_dropdown_validation(
    workbook: Any,
    validation_class: Any,
    *,
    sheet_name: str,
    column_name: str,
    values: list[str],
    max_rows: int = 1000,
) -> None:
    if sheet_name not in workbook.sheetnames:
        return
    worksheet = workbook[sheet_name]
    headers = [str(cell.value) for cell in worksheet[1]]
    if column_name not in headers:
        return
    column_index = headers.index(column_name) + 1
    column_letter = get_column_letter_for_index(column_index)
    formula = '"' + ",".join(values) + '"'
    if len(formula) > 255:
        helper_sheet_name = "_validation_lists"
        helper = (
            workbook[helper_sheet_name]
            if helper_sheet_name in workbook.sheetnames
            else workbook.create_sheet(helper_sheet_name)
        )
        helper.sheet_state = "hidden"
        helper_column_index = helper.max_column + 1
        if (
            helper.max_column == 1
            and helper.max_row == 1
            and helper.cell(row=1, column=1).value is None
        ):
            helper_column_index = 1
        helper_column_letter = get_column_letter_for_index(helper_column_index)
        helper.cell(row=1, column=helper_column_index).value = f"{sheet_name}.{column_name}"
        for row_index, value in enumerate(values, start=2):
            helper.cell(row=row_index, column=helper_column_index).value = value
        formula = (
            f"'{helper_sheet_name}'!${helper_column_letter}$2:"
            f"${helper_column_letter}${len(values) + 1}"
        )
    validation = validation_class(type="list", formula1=formula, allow_blank=True)
    validation.error = "Choose a supported Quant Studio value."
    validation.errorTitle = "Unsupported value"
    validation.prompt = "Use the dropdown or allowed_values sheet."
    validation.promptTitle = "Allowed values"
    worksheet.add_data_validation(validation)
    validation.add(f"{column_letter}2:{column_letter}{max_rows}")


def get_column_letter_for_index(column_index: int) -> str:
    """Thin wrapper to avoid importing openpyxl at module import time."""

    from openpyxl.utils import get_column_letter

    return get_column_letter(column_index)


def _apply_tab_colors(workbook: Any) -> None:
    tab_colors = {
        "instructions": "1F6EF5",
        "allowed_values": "607089",
        "transform_catalog": "607089",
        "examples": "607089",
        "required_columns": "607089",
        "schema": "0F8B5F",
        "feature_dictionary": "0F8B5F",
        "transformations": "D99A2B",
        "feature_review": "2A6F97",
        "scorecard_overrides": "C44536",
    }
    for sheet_name, color in tab_colors.items():
        if sheet_name in workbook.sheetnames:
            workbook[sheet_name].sheet_properties.tabColor = color


def default_challengers_for_target_mode(target_mode: TargetMode) -> list[ModelType]:
    """Provides sensible challenger defaults when no preset is selected."""

    if target_mode == TargetMode.BINARY:
        return [
            ModelType.ELASTIC_NET_LOGISTIC_REGRESSION,
            ModelType.SCORECARD_LOGISTIC_REGRESSION,
            ModelType.PROBIT_REGRESSION,
            ModelType.XGBOOST,
        ]
    if target_mode == TargetMode.MULTICLASS:
        return [
            ModelType.MULTINOMIAL_LOGISTIC_REGRESSION,
            ModelType.ORDINAL_LOGISTIC_REGRESSION,
            ModelType.DECISION_TREE,
        ]
    return [
        ModelType.BETA_REGRESSION,
        ModelType.TWO_STAGE_LGD_MODEL,
        ModelType.QUANTILE_REGRESSION,
        ModelType.XGBOOST,
    ]


MODEL_FAMILY_DEFINITIONS: dict[TargetMode, dict[str, tuple[str, tuple[ModelType, ...]]]] = {
    TargetMode.BINARY: {
        "binary_logistic_pd": (
            "Logistic / PD",
            (
                ModelType.LOGISTIC_REGRESSION,
                ModelType.ELASTIC_NET_LOGISTIC_REGRESSION,
                ModelType.SCORECARD_LOGISTIC_REGRESSION,
                ModelType.PROBIT_REGRESSION,
                ModelType.GEE_LOGISTIC_REGRESSION,
                ModelType.DISCRETE_TIME_HAZARD_MODEL,
            ),
        ),
        "binary_linear_baseline": (
            "Linear / Baseline",
            (ModelType.LINEAR_REGRESSION,),
        ),
        "binary_smooth_nonlinear": (
            "Smooth / Nonlinear",
            (
                ModelType.GAM_SPLINE_LOGISTIC,
                ModelType.DECISION_TREE,
                ModelType.RANDOM_FOREST,
                ModelType.EXTRA_TREES,
                ModelType.EXPLAINABLE_BOOSTING_MACHINE,
                ModelType.XGBOOST,
            ),
        ),
    },
    TargetMode.MULTICLASS: {
        "multiclass_class_models": (
            "Class Models",
            (
                ModelType.MULTINOMIAL_LOGISTIC_REGRESSION,
                ModelType.ORDINAL_LOGISTIC_REGRESSION,
                ModelType.DECISION_TREE,
            ),
        ),
    },
    TargetMode.CONTINUOUS: {
        "continuous_linear_regularized": (
            "Linear / Regularized",
            (
                ModelType.LINEAR_REGRESSION,
                ModelType.RIDGE_REGRESSION,
                ModelType.LASSO_REGRESSION,
                ModelType.ELASTIC_NET_REGRESSION,
            ),
        ),
        "continuous_lgd_severity": (
            "LGD / Severity",
            (
                ModelType.FRACTIONAL_LOGIT,
                ModelType.BETA_REGRESSION,
                ModelType.ZERO_ONE_INFLATED_BETA,
                ModelType.TWO_STAGE_LGD_MODEL,
                ModelType.TOBIT_REGRESSION,
            ),
        ),
        "continuous_glm_count_skewed": (
            "GLM / Count / Skewed",
            (
                ModelType.POISSON_REGRESSION,
                ModelType.NEGATIVE_BINOMIAL_REGRESSION,
                ModelType.GAMMA_REGRESSION,
                ModelType.TWEEDIE_REGRESSION,
            ),
        ),
        "continuous_panel_forecasting": (
            "Panel / Forecasting",
            (
                ModelType.PANEL_REGRESSION,
                ModelType.MIXED_EFFECTS_REGRESSION,
                ModelType.SARIMAX_FORECAST,
                ModelType.EXPONENTIAL_SMOOTHING_FORECAST,
                ModelType.UNOBSERVED_COMPONENTS_FORECAST,
            ),
        ),
        "continuous_survival_duration": (
            "Survival / Duration",
            (
                ModelType.COX_PROPORTIONAL_HAZARDS,
                ModelType.AFT_SURVIVAL_MODEL,
            ),
        ),
        "continuous_smooth_nonlinear": (
            "Smooth / Nonlinear",
            (
                ModelType.GAM_SPLINE_REGRESSION,
                ModelType.QUANTILE_REGRESSION,
                ModelType.DECISION_TREE,
                ModelType.RANDOM_FOREST,
                ModelType.EXTRA_TREES,
                ModelType.EXPLAINABLE_BOOSTING_MACHINE,
                ModelType.XGBOOST,
            ),
        ),
    },
}


def model_family_options_for_target_mode(
    target_mode: TargetMode,
    *,
    allowed_model_types: set[ModelType] | None = None,
) -> list[str]:
    """Returns selectable model-family keys for the selected target mode."""

    family_keys: list[str] = []
    for family_key, (_, model_types) in MODEL_FAMILY_DEFINITIONS[target_mode].items():
        if allowed_model_types is None or allowed_model_types.intersection(model_types):
            family_keys.append(family_key)
    return family_keys


def format_model_family(value: str) -> str:
    """Formats model-family keys for display."""

    for family_map in MODEL_FAMILY_DEFINITIONS.values():
        if value in family_map:
            return family_map[value][0]
    return value.replace("_", " ").title()


def model_family_for_model_type(
    model_type: ModelType,
    target_mode: TargetMode,
    *,
    allowed_model_types: set[ModelType] | None = None,
) -> str:
    """Finds the first family that contains a model type for a target mode."""

    for family_key, (_, model_types) in MODEL_FAMILY_DEFINITIONS[target_mode].items():
        if model_type in model_types and (
            allowed_model_types is None or model_type in allowed_model_types
        ):
            return family_key
    return model_family_options_for_target_mode(
        target_mode,
        allowed_model_types=allowed_model_types,
    )[0]


def model_types_for_family(
    target_mode: TargetMode,
    family_key: str,
    *,
    allowed_model_types: set[ModelType] | None = None,
) -> list[ModelType]:
    """Returns valid model types for a selected target-mode/family pair."""

    family_map = MODEL_FAMILY_DEFINITIONS[target_mode]
    if family_key not in family_map:
        family_key = model_family_options_for_target_mode(
            target_mode,
            allowed_model_types=allowed_model_types,
        )[0]
    valid_model_types = set(model_types_for_target_mode(target_mode))
    if allowed_model_types is not None:
        valid_model_types = valid_model_types.intersection(allowed_model_types)
    return [
        model_type
        for model_type in family_map[family_key][1]
        if model_type in valid_model_types
    ]


def model_types_for_target_mode(target_mode: TargetMode) -> list[ModelType]:
    """Returns model families that pass config validation for a target mode."""

    supported: list[ModelType] = []
    for model_type in ModelType:
        try:
            ModelConfig(model_type=model_type).validate(target_mode)
        except ValueError:
            continue
        supported.append(model_type)
    return supported


def subset_search_model_types_for_target_mode(target_mode: TargetMode) -> list[ModelType]:
    """Returns model families that support feature-subset search for a target mode."""

    return list(feature_subset_search_model_types_for_target_mode(target_mode))
