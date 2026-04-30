"""Regression coverage for the expanded governed transformation catalog."""

from __future__ import annotations

from io import BytesIO

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from quant_pd_framework import (
    ArtifactConfig,
    CleaningConfig,
    DataStructure,
    FeatureEngineeringConfig,
    FrameworkConfig,
    ModelConfig,
    ModelType,
    QuantModelOrchestrator,
    SplitConfig,
    TargetConfig,
    TargetMode,
    TransformationConfig,
    TransformationSpec,
    TransformationType,
)
from quant_pd_framework.gui_support import (
    build_column_editor_frame,
    build_feature_dictionary_editor_frame,
    build_feature_review_editor_frame,
    build_scorecard_override_editor_frame,
    build_template_workbook_bytes,
    build_transformation_editor_frame,
)
from tests.support import (
    build_binary_dataframe,
    build_common_schema,
    build_panel_forecast_dataframe,
    temporary_artifact_root,
)


def test_expanded_numeric_binning_encoding_date_and_missingness_transforms() -> None:
    dataframe = build_binary_dataframe(row_count=220)
    dataframe.loc[dataframe.index[::13], "recent_inquiries"] = np.nan

    with temporary_artifact_root("transformation_expansion_binary") as output_root:
        config = FrameworkConfig(
            schema=build_common_schema("account_id"),
            cleaning=CleaningConfig(),
            feature_engineering=FeatureEngineeringConfig(),
            target=TargetConfig(source_column="default_status", positive_values=[1]),
            split=SplitConfig(
                data_structure=DataStructure.TIME_SERIES,
                date_column="as_of_date",
                train_size=0.6,
                validation_size=0.2,
                test_size=0.2,
            ),
            model=ModelConfig(model_type=ModelType.LOGISTIC_REGRESSION),
            transformations=TransformationConfig(
                enabled=True,
                transformations=[
                    TransformationSpec(
                        transform_type=TransformationType.STANDARD_SCALE,
                        source_feature="balance",
                    ),
                    TransformationSpec(
                        transform_type=TransformationType.SAFE_RATIO,
                        source_feature="balance",
                        secondary_feature="tenure_months",
                    ),
                    TransformationSpec(
                        transform_type=TransformationType.QUANTILE_BINS,
                        source_feature="utilization",
                        parameter_value=4,
                    ),
                    TransformationSpec(
                        transform_type=TransformationType.WOE_ENCODING,
                        source_feature="utilization",
                        parameter_value=4,
                    ),
                    TransformationSpec(
                        transform_type=TransformationType.RARE_CATEGORY_COLLAPSE,
                        source_feature="channel",
                        parameter_value=0.1,
                    ),
                    TransformationSpec(
                        transform_type=TransformationType.FREQUENCY_ENCODING,
                        source_feature="channel",
                    ),
                    TransformationSpec(
                        transform_type=TransformationType.TARGET_ENCODING,
                        source_feature="channel",
                    ),
                    TransformationSpec(
                        transform_type=TransformationType.DATE_MONTH,
                        source_feature="as_of_date",
                    ),
                    TransformationSpec(
                        transform_type=TransformationType.ROW_MISSING_COUNT,
                        source_feature="balance, utilization, recent_inquiries",
                        output_feature="financial_missing_count",
                    ),
                ],
            ),
            artifacts=ArtifactConfig(output_root=output_root),
        )
        context = QuantModelOrchestrator(config=config).run(dataframe)

    expected_outputs = {
        "balance_standard_scaled",
        "balance_safe_over_tenure_months",
        "utilization_qbin",
        "utilization_woe",
        "channel_rare_collapsed",
        "channel_frequency",
        "channel_target_encoded",
        "as_of_date_month",
        "financial_missing_count",
    }
    assert expected_outputs.issubset(set(context.feature_columns))
    assert "governed_transformations" in context.diagnostics_tables
    assert context.working_data is not None
    assert "financial_missing_count" in context.working_data.columns


def test_expanded_temporal_transforms_respect_panel_ordering() -> None:
    dataframe = build_panel_forecast_dataframe(entity_count=8, periods_per_entity=12)
    dataframe["stress_event"] = dataframe["delinquency_rate"].gt(
        dataframe["delinquency_rate"].median()
    ).astype(int)

    with temporary_artifact_root("transformation_expansion_panel") as output_root:
        config = FrameworkConfig(
            schema=build_common_schema("segment_id"),
            cleaning=CleaningConfig(),
            feature_engineering=FeatureEngineeringConfig(),
            target=TargetConfig(
                source_column="forecast_value",
                mode=TargetMode.CONTINUOUS,
                output_column="forecast_value",
            ),
            split=SplitConfig(
                data_structure=DataStructure.PANEL,
                date_column="as_of_date",
                entity_column="segment_id",
                train_size=0.6,
                validation_size=0.2,
                test_size=0.2,
            ),
            model=ModelConfig(model_type=ModelType.PANEL_REGRESSION),
            transformations=TransformationConfig(
                enabled=True,
                transformations=[
                    TransformationSpec(
                        transform_type=TransformationType.ROLLING_SUM,
                        source_feature="utilization",
                        window_size=3,
                    ),
                    TransformationSpec(
                        transform_type=TransformationType.ROLLING_SLOPE,
                        source_feature="gdp_gap",
                        window_size=4,
                    ),
                    TransformationSpec(
                        transform_type=TransformationType.EXPANDING_MEAN,
                        source_feature="unemployment_rate",
                    ),
                    TransformationSpec(
                        transform_type=TransformationType.CUMULATIVE_SUM,
                        source_feature="stress_event",
                    ),
                    TransformationSpec(
                        transform_type=TransformationType.MONTHS_SINCE_EVENT,
                        source_feature="stress_event",
                    ),
                    TransformationSpec(
                        transform_type=TransformationType.CHANGE_FROM_BASELINE,
                        source_feature="delinquency_rate",
                    ),
                ],
            ),
            artifacts=ArtifactConfig(output_root=output_root),
        )

        context = QuantModelOrchestrator(config=config).run(dataframe)

    expected_outputs = {
        "utilization_rollsum_3",
        "gdp_gap_rollslope_4",
        "unemployment_rate_expanding_mean",
        "stress_event_cumulative_sum",
        "stress_event_months_since_event",
        "delinquency_rate_change_from_baseline",
    }
    assert expected_outputs.issubset(set(context.feature_columns))
    assert context.working_data is not None
    assert context.working_data.loc[:, list(expected_outputs)].notna().any().all()


def test_review_workbook_includes_transform_catalog_and_long_dropdown_validation() -> None:
    dataframe = build_binary_dataframe(row_count=20)
    workbook_bytes = build_template_workbook_bytes(
        schema_frame=build_column_editor_frame(dataframe),
        feature_dictionary_frame=build_feature_dictionary_editor_frame(dataframe),
        transformation_frame=build_transformation_editor_frame(),
        feature_review_frame=build_feature_review_editor_frame(),
        scorecard_override_frame=build_scorecard_override_editor_frame(),
    )

    workbook_sheets = pd.read_excel(BytesIO(workbook_bytes), sheet_name=None)
    assert "transform_catalog" in workbook_sheets
    assert TransformationType.WOE_ENCODING.value in set(
        workbook_sheets["transform_catalog"]["transform_type"]
    )

    workbook = load_workbook(BytesIO(workbook_bytes))
    assert "_validation_lists" in workbook.sheetnames
    assert workbook["_validation_lists"].sheet_state == "hidden"
    validations = workbook["transformations"].data_validations.dataValidation
    assert any(validation.type == "list" for validation in validations)
